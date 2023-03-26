__author__     = "Nandan Sharma"
__title__      = "T3000 Python Scripts"

from tkinter import Tk, StringVar, Label, Entry, Button, filedialog, Frame, Radiobutton, Menu
from utils import p1_parameterChecker as p1
from utils import p2_fupScheduleChecker as p2
from utils import p3_fupCompare as p3
import sys
from os import path, startfile, mkdir
from xlrd import open_workbook as open_wb
from xlutils.copy import copy as copy_wb
from xlwt import easyxf
from shutil import copyfile
from openpyxl import Workbook as openpyxl_wb
from re import sub
from webbrowser import open_new_tab
import json

def to_be_defined_later():
    pass
    
# determine if application is a script file or frozen exe
if getattr(sys, 'frozen', False):
    application_path = path.dirname(path.realpath(sys.executable))
elif __file__:
    application_path = path.dirname(__file__)
    
    
initDir = path.expanduser('~')
p1help_file = path.join(application_path, 'help', 'Parameter checker.xhtml')
p2help_file = path.join(application_path, 'help', 'FuP schedule checker.xhtml')


# make folder "result"
if not path.exists(path.join(application_path, 'result')):
    mkdir(path.join(application_path, 'result'))
    
    
# p1 variables
p1spec_file = ""
p1spec_sh_name = "ParameterKeys"
p1t3k_exp_file = ""
p1t3k_imp_file = path.join(application_path, 'result', 'T3000_param_import.xls')
p1matched_param_file = path.join(application_path, 'result', 'matched_param.xlsx')
p1unmatched_param_t3k_file = path.join(application_path, 'result', 'unmatched_param_T3000.xlsx')
p1unmatched_param_spec_file = path.join(application_path, 'result', 'unmatched_param_spec.xlsx')
        
# p2 variables      
p2stdSchFile = ""
p2projSchFileOld = ""
p2projSchFileNew = path.join(application_path, 'result', 'T3000_schedule_import.xls')
p2projSchFileNewChk = path.join(application_path, 'result', 'T3000_schedule_check.xls')

# p3 variables      
p3oldJarExtractDir = ""
p3newJarExtractDir = ""
p3diff_file = path.join(application_path, 'result', 'T3000_jar_extract_diff.xlsx')
T3000_AF_db_dict_json = path.join(application_path, 'utils', 'T3000_AF_db_dict.json')
with open(T3000_AF_db_dict_json) as json_file:
    T3000_AF_db_dict = json.load(json_file)

# p1 function definitions        
# Functions for opening the file explorer windows

def p1browseFiles_spec():
    
    global p1spec_file
    global p1file_explore_spec_label
    p1spec_file = filedialog.askopenfilename(initialdir = initDir,
                                          title = "Select the specification file",
                                          filetypes = (("xls files",
                                                        "*.xls"),
                                                       ("all files",
                                                        "*.*")))
    p1file_explore_spec_label.configure(text = p1spec_file)

    workbook = open_wb(p1spec_file)
    sheet = workbook.sheet_by_name(p1spec_sh_name)
    col_count = sheet.ncols

    for cur_col in range(0, col_count):
        var = str(sheet.cell(0, cur_col))
        var = sub('^text:', '', var)
        var = sub("'","", var) # remove all single quote chars from the value
        var = sub('"','', var) # remove all double quote chars from the value
        var = sub(' ','', var) # remove all space chars from the value
        var = sub('   ','', var) # remove all tab chars from the value
        if 'parameterkey' == var.lower():
            p1e4_var.set(p1.num2col(cur_col))
            break
        elif cur_col == col_count - 1:
            for cur_col in range(1, col_count):
                var = str(sheet.cell(1, cur_col))
                var = sub('^text:', '', var)
                var = sub("'","", var) # remove all single quote chars from the value
                var = sub('"','', var) # remove all double quote chars from the value
                var = sub(' ','', var) # remove all space chars from the value
                var = sub('   ','', var) # remove all tab chars from the value
                if 'parameterkey' == var.lower():
                    p1e4_var.set(p1.num2col(cur_col))
                    break
        else:
            p1e4_var.set('Fill?')

    for cur_col in range(0, col_count):
        var = str(sheet.cell(0, cur_col))
        var = sub('^text:', '', var)
        var = sub("'","", var) # remove all single quote chars from the value
        var = sub('"','', var) # remove all double quote chars from the value
        var = sub(' ','', var) # remove all space chars from the value
        var = sub('   ','', var) # remove all tab chars from the value
        if 'value' == var.lower():
            p1e5_var.set(p1.num2col(cur_col))
            break
        elif cur_col == col_count - 1:
            for cur_col in range(1, col_count):
                var = str(sheet.cell(1, cur_col))
                var = sub('^text:', '', var)
                var = sub("'","", var) # remove all single quote chars from the value
                var = sub('"','', var) # remove all double quote chars from the value
                var = sub(' ','', var) # remove all space chars from the value
                var = sub('   ','', var) # remove all tab chars from the value
                if 'value' == var.lower():
                    p1e5_var.set(p1.num2col(cur_col))
                    break
        else:
            p1e5_var.set('Fill?')
        

def p1browseFiles_t3k():
    
    global p1t3k_exp_file
    global p1file_explore_t3k_label
    p1t3k_exp_file = filedialog.askopenfilename(initialdir = initDir,
                                          title = "Select the t3k export file",
                                          filetypes = (("xls files",
                                                        "*.xls"),
                                                       ("all files",
                                                        "*.*")))
    p1file_explore_t3k_label.configure(text = p1t3k_exp_file)

    workbook = open_wb(p1t3k_exp_file)
    sheet = workbook.sheet_by_index(0)
    col_count = sheet.ncols

    for cur_col in range(0, col_count):
        var = str(sheet.cell(0, cur_col))
        var = sub('^text:', '', var)
        var = sub("'","", var) # remove all single quote chars from the value
        var = sub('"','', var) # remove all double quote chars from the value
        var = sub(' ','', var) # remove all space chars from the value
        var = sub('   ','', var) # remove all tab chars from the value
        if 'parameterkey' == var.lower():
            p1e6_var.set(p1.num2col(cur_col))
            break
        else:
            p1e6_var.set('Fill?')

    for cur_col in range(0, col_count):
        var = str(sheet.cell(0, cur_col))
        var = sub('^text:', '', var)
        var = sub("'","", var) # remove all single quote chars from the value
        var = sub('"','', var) # remove all double quote chars from the value
        var = sub(' ','', var) # remove all space chars from the value
        var = sub('   ','', var) # remove all tab chars from the value
        if 'value' == var.lower():
            p1e7_var.set(p1.num2col(cur_col))
            break
        else:
            p1e7_var.set('Fill?')


# define generate_files function
def p1generate_files():
    
    global p1gen_file_label

    try:
        # set user inputs
        prefix = p1e1_var.get()
        dev_th_pct = p1e2_var.get()
        incl_dev_th = p1e3_var.get()
        spec_col_key = p1e4_var.get()
        spec_col_val = p1e5_var.get()
        t3k_col_key = p1e6_var.get()
        t3k_col_val = p1e7_var.get()

        dev_th_pct = float(dev_th_pct)
        incl_dev_th = incl_dev_th.lower() in ('yes', 'true', 't', '1')

        spec_dict = p1.create_spec_dict(p1spec_file, spec_col_key, spec_col_val, p1spec_sh_name)

        t3k_list_key, t3k_list_val = p1.create_t3k_list(p1t3k_exp_file, prefix, t3k_col_key, t3k_col_val)

        t3k_list_val_updated, matched_param_list, unmatched_param_list_t3k, unmatched_param_list_spec = p1.update_t3k_list(spec_dict,
                                                                                                                        t3k_list_key,
                                                                                                                        t3k_list_val,
                                                                                                                        incl_dev_th,
                                                                                                                        dev_th_pct)
        # write results
        copyfile(p1t3k_exp_file, p1t3k_imp_file)
        rb = open_wb(p1t3k_imp_file, formatting_info = True)
        wb = copy_wb(rb)
        s = wb.get_sheet(0)
        style = easyxf('pattern: pattern solid, fore_colour yellow')

        for idx, key in enumerate(t3k_list_key):
            if t3k_list_val_updated[idx] != t3k_list_val[idx]:
                s.write(idx + 1, p1.col2num(t3k_col_val), t3k_list_val_updated[idx], style)

        wb.save(p1t3k_imp_file)

        wb1 = openpyxl_wb()
        ws = wb1.active
        for row in matched_param_list: ws.append(row)
        wb1.save(p1matched_param_file)

        wb2 = openpyxl_wb()
        ws = wb2.active
        for row in unmatched_param_list_t3k: ws.append(row)
        wb2.save(p1unmatched_param_t3k_file)

        wb3 = openpyxl_wb()
        ws = wb3.active
        for row in unmatched_param_list_spec: ws.append(row)
        wb3.save(p1unmatched_param_spec_file)

        startfile(path.join(application_path, 'result'))

        p1gen_file_label.configure(text = 'Success', fg = "green")

    except PermissionError:
        p1gen_file_label.configure(text = 'Permission to read/write file denied', fg = "red")
    except:
        p1gen_file_label.configure(text = 'Failure', fg = "red")
        
        
def call_p1():
    global widget_pack_list
    for widget in widget_pack_list:
        widget.pack_forget()
    
    title_widget = [p1frame_title, p1title_label]

    input_widget = [p1frame_input, p1e1_label, p1e1_entry, p1e2_label, p1e2_entry, p1e3_label, p1e3_entry, p1e4_label, p1e4_entry, 
                    p1e5_label, p1e5_entry, p1e6_label, p1e6_entry, p1e7_label, p1e7_entry]
                              
    btn_widget = [p1frame_btn, p1gen_file_label, p1gen_file_btn, p1file_explore_spec_label, p1explore_spec_btn, p1file_explore_t3k_label, p1explore_t3k_btn]

    widget_pack_list = title_widget + input_widget + btn_widget
    
    for widget in widget_pack_list:
        if type(widget) == Frame: widget.pack(pady = 5, side = 'top')
        else: widget.pack(pady = 1, side = 'top')
        
        
# p2 function definitions        
# Functions for opening the file explorer windows

def p2browseFiles_std():

    global p2stdSchFile
    global p2file_explore_std_label
    p2stdSchFile = filedialog.askopenfilename(initialdir = initDir,
                                          title = "Select the standard schedule table",
                                          filetypes = (("xls files",
                                                        "*.xls"),
                                                       ("all files",
                                                        "*.*")))
    p2file_explore_std_label.configure(text = p2stdSchFile)
    
    workbook = open_wb(p2stdSchFile)
    sheet = workbook.sheet_by_index(0)
    
    var = str(sheet.cell(1, 1))
    var = sub('^text:', '', var)
    var = sub("'","", var) # remove all single quote chars from the value
    var = sub('"','', var) # remove all double quote chars from the value
    var = sub(' ','', var) # remove all space chars from the value
    var = sub('   ','', var) # remove all tab chars from the value
    var = var[0:2]
    p2e1_var.set(var)
    
    
def p2browseFiles_prj():

    global p2projSchFileOld
    global p2file_explore_prj_label
    p2projSchFileOld = filedialog.askopenfilename(initialdir = initDir,
                                          title = "Select the project schedule table",
                                          filetypes = (("xls files",
                                                        "*.xls"),
                                                       ("all files",
                                                        "*.*")))
    p2file_explore_prj_label.configure(text = p2projSchFileOld)
    
    workbook = open_wb(p2projSchFileOld)
    sheet = workbook.sheet_by_index(0)
    
    var = str(sheet.cell(1, 1))
    var = sub('^text:', '', var)
    var = sub("'","", var) # remove all single quote chars from the value
    var = sub('"','', var) # remove all double quote chars from the value
    var = sub(' ','', var) # remove all space chars from the value
    var = sub('   ','', var) # remove all tab chars from the value
    var = var[0:2]
    p2e2_var.set(var)

    var = str(sheet.cell(1, 2))
    var = sub('^text:', '', var)
    var = sub("'","", var) # remove all single quote chars from the value
    var = sub('"','', var) # remove all double quote chars from the value
    var = sub(' ','', var) # remove all space chars from the value
    var = sub('   ','', var) # remove all tab chars from the value
    var = var.split("/")
    if "AP1" in var: p2e3_var.set("AP1")
    elif "AP2" in var: p2e3_var.set("AP2")
    elif "AP3" in var: p2e3_var.set("AP3")
    elif "AP4" in var: p2e3_var.set("AP4")
    elif "AP5" in var: p2e3_var.set("AP5")
    elif "AP6" in var: p2e3_var.set("AP6")
    elif "AP7" in var: p2e3_var.set("AP7")
    elif "AP8" in var: p2e3_var.set("AP8")
    elif "AP9" in var: p2e3_var.set("AP9")
    elif "AP10" in var: p2e3_var.set("AP10")
    else: p2e3_var.set("Fill?")


# define generate_files function
def p2generate_files():

    global p2gen_file_label

    try:
        # set user inputs
        prefix_std = p2e1_var.get()
        prefix_prj = p2e2_var.get()
        ap = p2e3_var.get()
        ap = "/" + ap + "/"

        stdSchList = p2.create_std_sch_list(p2stdSchFile, ap, prefix_std, prefix_prj)
        assert stdSchList != []
        projSchOldList = p2.create_proj_sch_old_list(p2projSchFileOld, ap)
        projSchNewList = p2.create_proj_sch_new_list(stdSchList, projSchOldList)
        
        # write results
        copyfile(p2projSchFileOld, p2projSchFileNewChk)
        rb = open_wb(p2projSchFileNewChk, formatting_info = True)
        wb = copy_wb(rb)
        s = wb.get_sheet(0)
        style = easyxf('pattern: pattern solid, fore_colour yellow')
        
        s.write(0, 6, "Remarks")
        s.write(0, 7, "Standard Seq")

        for idx_row, subList in enumerate(projSchNewList):
            if subList[-2] == "check":
                s.write(idx_row + 1, 0, idx_row + 1, style)
                for idx_col in range(1, len(subList)):
                    s.write(idx_row + 1, idx_col, subList[idx_col], style)
            else:
                s.write(idx_row + 1, 0, idx_row + 1)
                for idx_col in range(1, len(subList)):
                    s.write(idx_row + 1, idx_col, subList[idx_col])
                
        wb.save(p2projSchFileNewChk)
        
        copyfile(p2projSchFileOld, p2projSchFileNew)
        rb = open_wb(p2projSchFileNew, formatting_info = True)
        wb = copy_wb(rb)
        s = wb.get_sheet(0)

        for idx_row, subList in enumerate(projSchNewList):
            if subList[-2] == "check":
                s.write(idx_row + 1, 0, idx_row + 1, style)
                for idx_col in range(1, len(subList) - 2): # minus 2 for skipping last two rows.
                    s.write(idx_row + 1, idx_col, subList[idx_col], style)
            else:
                s.write(idx_row + 1, 0, idx_row + 1)
                for idx_col in range(1, len(subList) - 2): # minus 2 for skipping last two rows.
                    s.write(idx_row + 1, idx_col, subList[idx_col])

        wb.save(p2projSchFileNew)

        startfile(path.join(application_path, 'result'))

        p2gen_file_label.configure(text = 'Success', fg = "green")
    except AssertionError:
        p2gen_file_label.configure(text = 'Restricted access to standard schedule file', fg = "red")
    except PermissionError:
        p2gen_file_label.configure(text = 'Permission to read/write file denied', fg = "red")
    except:
        p2gen_file_label.configure(text = 'Failure', fg = "red")
        
        
def call_p2():
    global widget_pack_list
    for widget in widget_pack_list:
        widget.pack_forget()
    
    title_widget = [p2frame_title, p2title_label]

    input_widget = [p2frame_input, p2e1_label, p2e1_entry, p2e2_label, p2e2_entry, p2e3_label, p2e3_entry]
                              
    btn_widget = [p2frame_btn, p2gen_file_label, p2gen_file_btn, p2file_explore_std_label, p2explore_std_btn, p2file_explore_prj_label, p2explore_prj_btn]

    widget_pack_list = title_widget + input_widget + btn_widget

    for widget in widget_pack_list:
        if type(widget) == Frame: widget.pack(pady = 5, side = 'top')
        else: widget.pack(pady = 1, side = 'top')


# p3 function definitions        
# Functions for opening the folder explorer windows

def p3browseDir_oldJarExtract():

    global p3oldJarExtractDir
    global p3dir_explore_oldJarExtract_label
    p3oldJarExtractDir = filedialog.askdirectory(initialdir = initDir,
                                          title = "Select the old jar extract folder")
    p3dir_explore_oldJarExtract_label.configure(text = p3oldJarExtractDir)
    
def p3browseDir_newJarExtract():

    global p3newJarExtractDir
    global p3dir_explore_newJarExtract_label
    p3newJarExtractDir = filedialog.askdirectory(initialdir = initDir,
                                          title = "Select the new jar extract folder")
    p3dir_explore_newJarExtract_label.configure(text = p3newJarExtractDir)

# define generate_files function
def p3generate_files():
    
    global p3gen_file_label
    global T3000_AF_db_dict

    try:
        # set user inputs
        ignore_inp_sig_desig = p3e1_var.get()
        ignore_inp_sig_desig = ignore_inp_sig_desig.lower() in ('yes', 'true', 't', '1')

        p3.diff_fup_dir(p3oldJarExtractDir, p3newJarExtractDir, ignore_inp_sig_desig, p3diff_file, T3000_AF_db_dict, None, None)

        startfile(path.join(application_path, 'result'))

        p3gen_file_label.configure(text = 'Success', fg = "green")

    except:
        p3gen_file_label.configure(text = 'Failure', fg = "red")
    
    # finally:
        # # set user inputs
        # ignore_inp_sig_desig = p3e1_var.get()
        # ignore_inp_sig_desig = ignore_inp_sig_desig.lower() in ('yes', 'true', 't', '1')

        # p3.diff_fup_dir(p3oldJarExtractDir, p3newJarExtractDir, ignore_inp_sig_desig, p3diff_file, T3000_AF_db_dict, None, None)

        # startfile(path.join(application_path, 'result'))

        # p3gen_file_label.configure(text = 'Success', fg = "green")

def call_p3():
    global widget_pack_list
    for widget in widget_pack_list:
        widget.pack_forget()
    
    title_widget = [p3frame_title, p3title_label]

    input_widget = [p3frame_input, p3e1_label, p3e1_entry]
                              
    btn_widget = [p3frame_btn, p3gen_file_label, p3gen_file_btn, p3dir_explore_oldJarExtract_label, p3dir_explore_oldJarExtract_btn, p3dir_explore_newJarExtract_label, p3dir_explore_newJarExtract_btn]

    widget_pack_list = title_widget + input_widget + btn_widget
    
    for widget in widget_pack_list:
        if type(widget) == Frame: widget.pack(pady = 5, side = 'top')
        else: widget.pack(pady = 1, side = 'top')


root = Tk()

# set window size and title
root.geometry("600x600")
root.title(__title__)


menuRoot = Menu(root)
root.config(menu=menuRoot)

selectProgramMenu = Menu(menuRoot, tearoff = 0)
menuRoot.add_cascade(label="Select Program", menu=selectProgramMenu)
selectProgramMenu.add_command(label="Parameter checker", command = call_p1)
selectProgramMenu.add_command(label="FuP schedule checker", command = call_p2)
selectProgramMenu.add_command(label="Jar extract comparator", command = call_p3)

helpMenu = Menu(menuRoot, tearoff = 0)
menuRoot.add_cascade(label="User Guide", menu=helpMenu)
helpMenu.add_command(label="Parameter checker", command = lambda: open_new_tab(p1help_file))
helpMenu.add_command(label="FuP schedule checker", command = lambda: open_new_tab(p2help_file))
helpMenu.add_command(label="Jar extract comparator", command = to_be_defined_later)


author_frame = Frame(root)
author_label = Label(author_frame, text = "Author: " + __author__, font=('calibre',8, 'normal'), foreground = "gray75")
author_frame.pack(fill='both', side='bottom')
author_label.pack(side='right')


# p1 widgets
# declare string variable with default value for widgets
p1e1_var = StringVar(root, value='COMOS')
p1e2_var = StringVar(root, value='10')
p1e3_var = StringVar(root, value='YES')
p1e4_var = StringVar(root)
p1e5_var = StringVar(root)
p1e6_var = StringVar(root)
p1e7_var = StringVar(root)

# title label
p1frame_title = Frame(root)
p1title_label = Label(p1frame_title, text = 'Parameter checker', font=('calibre',15, 'normal'), foreground = 'light blue')

# create labels for user inputs
p1frame_input = Frame(root)
p1e1_label = Label(p1frame_input, text = 'Prefix for parameter key identification in T3000 (case sensitive)', font=('calibre',10, 'bold'))
p1e2_label = Label(p1frame_input, text = 'Allowed deviation threshold in %', font=('calibre',10, 'bold'))
p1e3_label = Label(p1frame_input, text = 'Include parameter values with deviation higher than the threshold (YES/NO)', font=('calibre',10, 'bold'))
p1e4_label = Label(p1frame_input, text = 'Key column in specification file', font=('calibre',10, 'bold'))
p1e5_label = Label(p1frame_input, text = 'Value column in specification file', font=('calibre',10, 'bold'))
p1e6_label = Label(p1frame_input, text = 'Key column in T3000 export file', font=('calibre',10, 'bold'))
p1e7_label = Label(p1frame_input, text = 'Value column in T3000 export file', font=('calibre',10, 'bold'))

# create entries for user inputs
p1e1_entry = Entry(p1frame_input, textvariable = p1e1_var, font=('calibre',10,'normal'))
p1e2_entry = Entry(p1frame_input, textvariable = p1e2_var, font=('calibre',10,'normal'))
p1e3_entry = Entry(p1frame_input, textvariable = p1e3_var, font=('calibre',10,'normal'))
p1e4_entry = Entry(p1frame_input, textvariable = p1e4_var, font=('calibre',10,'normal'))
p1e5_entry = Entry(p1frame_input, textvariable = p1e5_var, font=('calibre',10,'normal'))
p1e6_entry = Entry(p1frame_input, textvariable = p1e6_var, font=('calibre',10,'normal'))
p1e7_entry = Entry(p1frame_input, textvariable = p1e7_var, font=('calibre',10,'normal'))

# create Button that will call the function generate_files
p1frame_btn = Frame(root)
p1gen_file_label = Label(p1frame_btn, text = '', font=('calibre',10, 'bold'), fg = "green")
p1gen_file_btn = Button(p1frame_btn,text = 'Generate import file', command = p1generate_files)

# Create File Explorers
p1file_explore_spec_label = Label(p1frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p1explore_spec_btn = Button(p1frame_btn, text = "Browse specification file", command = p1browseFiles_spec)
p1file_explore_t3k_label = Label(p1frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p1explore_t3k_btn = Button(p1frame_btn, text = "Browse T3000 export file", command = p1browseFiles_t3k)


# p2 widgets
# declare string variable with default value for widgets
p2e1_var = StringVar(root) # prefix_std
p2e2_var = StringVar(root) # prefix_prj
p2e3_var = StringVar(root) # ap

# title label
p2frame_title = Frame(root)
p2title_label = Label(p2frame_title, text = 'FuP schedule checker', font=('calibre',15, 'normal'), foreground = 'light blue')

# create labels for user inputs
p2frame_input = Frame(root)
p2e1_label = Label(p2frame_input, text = 'Standard KKS prefix', font=('calibre',10, 'bold'))
p2e2_label = Label(p2frame_input, text = 'Project KKS prefix', font=('calibre',10, 'bold'))
p2e3_label = Label(p2frame_input, text = 'AP name', font=('calibre',10, 'bold'))

# create entries for user inputs
p2e1_entry = Entry(p2frame_input, textvariable = p2e1_var, font=('calibre',10,'normal'))
p2e2_entry = Entry(p2frame_input, textvariable = p2e2_var, font=('calibre',10,'normal'))
p2e3_entry = Entry(p2frame_input, textvariable = p2e3_var, font=('calibre',10,'normal'))

# create Button that will call the function generate_files
p2frame_btn = Frame(root)
p2gen_file_label = Label(p2frame_btn, text = '', font=('calibre',10, 'bold'), fg = "green")
p2gen_file_btn = Button(p2frame_btn,text = 'Generate import file', command = p2generate_files)

# Create File Explorers
p2file_explore_std_label = Label(p2frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p2explore_std_btn = Button(p2frame_btn, text = "Browse standard schedule table", command = p2browseFiles_std)
p2file_explore_prj_label = Label(p2frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p2explore_prj_btn = Button(p2frame_btn, text = "Browse project schedule table", command = p2browseFiles_prj)


# p3 widgets
# declare string variable with default value for widgets
p3e1_var = StringVar(root, value='NO') # ignore_inp_sig_desig

# title label
p3frame_title = Frame(root)
p3title_label = Label(p3frame_title, text = 'Jar extract comparator', font=('calibre',15, 'normal'), foreground = 'light blue')

# create labels for user inputs
p3frame_input = Frame(root)
p3e1_label = Label(p3frame_input, text = 'Ignore designation for input signals (YES/NO)', font=('calibre',10, 'bold'))

# create entries for user inputs
p3e1_entry = Entry(p3frame_input, textvariable = p3e1_var, font=('calibre',10,'normal'))

# create Button that will call the function generate_files
p3frame_btn = Frame(root)
p3gen_file_label = Label(p3frame_btn, text = '', font=('calibre',10, 'bold'), fg = "green")
p3gen_file_btn = Button(p3frame_btn,text = 'Generate difference file', command = p3generate_files)

# Create Folder Explorers
p3dir_explore_oldJarExtract_label = Label(p3frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p3dir_explore_oldJarExtract_btn = Button(p3frame_btn, text = "Browse old jar extract folder", command = p3browseDir_oldJarExtract)
p3dir_explore_newJarExtract_label = Label(p3frame_btn, text = '', font=('calibre',10, 'bold'), fg = "blue")
p3dir_explore_newJarExtract_btn = Button(p3frame_btn, text = "Browse new jar extract folder", command = p3browseDir_newJarExtract)


widget_pack_list = []
title_widget = [p1frame_title, p1title_label]

input_widget = [p1frame_input, p1e1_label, p1e1_entry, p1e2_label, p1e2_entry, p1e3_label, p1e3_entry, p1e4_label, p1e4_entry, 
                p1e5_label, p1e5_entry, p1e6_label, p1e6_entry, p1e7_label, p1e7_entry]
                          
btn_widget = [p1frame_btn, p1gen_file_label, p1gen_file_btn, p1file_explore_spec_label, p1explore_spec_btn, p1file_explore_t3k_label, p1explore_t3k_btn]

widget_pack_list = title_widget + input_widget + btn_widget

for widget in widget_pack_list:
    if type(widget) == Frame: widget.pack(pady = 5, side = 'top')
    else: widget.pack(pady = 1, side = 'top')


# perform an infinite loop for the window to display
root.mainloop()